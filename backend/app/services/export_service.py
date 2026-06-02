import csv
import io
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import DocumentLog, AckStatus
from app.timeutil import fmt_local

_HEADERS = [
    "#", "FISH", "Login", "Tarkibiy tuzilma", "Lavozim", "Email",
    "Holati", "Ochilgan vaqti", "Tanishib chiqqan vaqti", "IP manzil",
]

_STATUS_LABELS = {
    AckStatus.OPENED: "Hujjat ochilgan",
    AckStatus.ACKNOWLEDGED: "Tanishib chiqilgan",
}


def _fmt_dt(dt: datetime | None) -> str:
    return fmt_local(dt)


def _row(idx: int, log: DocumentLog) -> list:
    return [
        idx,
        log.full_name or "",
        log.ad_username or "",
        log.department or "",
        log.position or "",
        log.email or "",
        _STATUS_LABELS.get(log.status, log.status),
        _fmt_dt(log.opened_at),
        _fmt_dt(log.acknowledged_at),
        log.ip_address or "",
    ]


def export_csv(logs: List[DocumentLog]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_HEADERS)
    for idx, log in enumerate(logs, start=1):
        writer.writerow(_row(idx, log))
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat


def export_excel(logs: List[DocumentLog], doc_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acknowledgements"

    last_col = len(_HEADERS)
    last_letter = ws.cell(row=1, column=last_col).column_letter

    # Title row
    ws.merge_cells(f"A1:{last_letter}1")
    title_cell = ws["A1"]
    title_cell.value = doc_title
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        AckStatus.OPENED: "FFF2CC",
        AckStatus.ACKNOWLEDGED: "E2EFDA",
    }

    for row_idx, log in enumerate(logs, start=3):
        data = _row(row_idx - 2, log)
        row_fill = PatternFill("solid", fgColor=status_colors.get(log.status, "FFFFFF"))
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill

    col_widths = [5, 26, 18, 24, 22, 26, 20, 20, 22, 16]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
